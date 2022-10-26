#!python
from posixpath import split
import sys
import datetime
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# * Function


# Get the time string from the log
def get_time_str(str):
    return str[1:24]


# Convert time string to epoch time
def get_epoch_time(str):
    str = get_time_str(str)
    # print(str[0:4], str[5:7], str[8:10], str[11:13],
    #       str[14:16], str[17:19], str[20:24])
    time = datetime.datetime(int(str[0:4]),
                             int(str[5:7]),
                             int(str[8:10]),
                             int(str[11:13]),
                             int(str[14:16]),
                             int(str[17:19]),
                             int(str[20:24])*1000).timestamp()
    return time

def print_help():
    print(f"{__file__.split('/')[-1]} [<log file>] [-s|--start <string>] [-e|--end <string>]")


# TODO : this tool help info
# TODO : version
print("Reboot Druation : Ver 1.1")
# TODO : example

# TODO : python 2.7 compatibility
# TODO : Check python tool: matplot/numpy/datetime
# TODO : pip install help info

# Argument Check
if (len(sys.argv) < 2):
    print("Not enough argument")
    print_help()
    exit()

# * Constant
# Define Start str and End str in log
MSG_START = "] System enter S5 state"
MSG_END = "] BIOS boot up from SPI "
STATE_START = "SHUTDOWN/START"
STATE_END = "BOOT UP/END"

i = 1
while i<len(sys.argv):
    if sys.argv[i] == '-s' or sys.argv[i] == '--start':
        MSG_START = sys.argv[i+1]
        i = i + 2
    elif sys.argv[i] == '-e' or sys.argv[i] == '--end':
        MSG_END = sys.argv[i+1]
        i = i + 2
    else:
        name_log = sys.argv[i]
        i = i + 1


# * Open log file
if not(os.path.exists(name_log)):
    print(f"File \"{name_log}\" not found")
    print_help()
    exit()

data_log = open(name_log, 'rb').read().decode('ISO-8859-1')
data_log = data_log.split("\r\n")

# ! Main Function
array_shutdown_time = []    # Shutdown Time(s) with Zeroing
array_bootup_time = []      # Bootup Time(s) with Zeroing
array_event_time = []       # Event Time(s) with Zeroing
array_event_date = []       # Shutdown/Bootup date YY-MM-DD HH:mm:SS without Zeroing
array_event_state = []      # Event marker use with array_event_date
flag_record = False         # Only record the next bootup after shutdown

for index, line in enumerate(data_log):
    if MSG_START in line:
        array_event_time.append(get_epoch_time(line))
        array_shutdown_time.append(get_epoch_time(line))
        array_event_date.append(get_time_str(line))
        array_event_state.append(STATE_START)
        flag_record = True
    if MSG_END in line:
        if flag_record:
            array_event_time.append(get_epoch_time(line))
            array_bootup_time.append(get_epoch_time(line))
            array_event_date.append(get_time_str(line))
            array_event_state.append(STATE_END)
            flag_record = False

# Bootup and shutdown cycle not match
if len(array_bootup_time) != len(array_shutdown_time):
    array_shutdown_time.pop()

# Quit : if there are not cycle event in log
if len(array_bootup_time) == 0:
    sys.exit("No cycle event found in this log")

array_event_time = np.array(array_event_time)
array_bootup_time = np.array(array_bootup_time)
array_shutdown_time = np.array(array_shutdown_time)
# Zeroing : calculate from the beginning of log
array_event_time = array_event_time - array_shutdown_time[0]
array_bootup_time = array_bootup_time - array_shutdown_time[0]
array_shutdown_time = array_shutdown_time - array_shutdown_time[0]
result = array_bootup_time - array_shutdown_time

# * Plot
fig = plt.figure()
plt.title("Boot up duration for each Cycle")
plt.xlabel("Cycle count from the beginning of log")
plt.ylabel("Boot up duration (s)")
plt.plot(result, ".")
plt.savefig(f'{name_log}_result.png')
plt.savefig(f'{name_log}_result.svg')
plt.close()
fig = plt.figure()
count, bins, ignored = plt.hist(result, 30, density=False)
plt.gca().set(title='Frequency Histogram',
              xlabel='Boot up Duration(s)', ylabel='Count')
plt.savefig(f'{name_log}_result_distribution.png')
plt.savefig(f'{name_log}_result_distribution.svg')
# plt.show()


# * Data anal
print(f"Average duration : {np.average(result)}")
print(f"Standard deviation : {np.std(result)}")
# TODO : Off-peak value

# * Export excel
wb = Workbook()
ws_result = wb.active
ws_result.title = "Boot up Duration"
ws_all = wb.create_sheet("All Data")

# Assign Sheet-result
# Title and font
title_result = ["Cycle", "Boot up Duration(s)"]
ws_result.append(title_result)
ws_result.row_dimensions[1].font = Font(bold=True)
# Data
SIZE = len(result)
for r in range(SIZE):
    ws_result.cell(row=r+2, column=1, value=r+1)        # Cycle
    ws_result.cell(row=r+2, column=2, value=result[r])  # Duration(s)
char = get_column_letter(2)
# Avgrage
ws_result.cell(row=SIZE+3, column=1, value="AVG")
ws_result.cell(row=SIZE+3, column=2,
               value=f"=AVERAGE({char}2:{char}{SIZE+2})")
ws_result.cell(row=SIZE+4, column=1, value="STD")
ws_result.cell(row=SIZE+4, column=2,
               value=f"=STDEV({char}2:{char}{SIZE+2})")


# Assign Sheet-all
# Title and font
title_all = ["Date", "Event", "Cycle",
             "Time(s) : \n calculate from the beginning of the log", "Boot up duration(s)"]
ws_all.append(title_all)
ws_all.row_dimensions[1].font = Font(bold=True)

# Data
cycle = 0
SIZE = len(array_event_date)
for r in range(SIZE):
    ws_all.cell(row=r+2, column=1, value=array_event_date[r])   # Date
    ws_all.cell(row=r+2, column=2, value=array_event_state[r])  # Event
    ws_all.cell(row=r+2, column=4, value=array_event_time[r])   # Times(s)
    if array_event_state[r] == STATE_START:
        ws_all.cell(row=r+2, column=3, value=cycle+1)           # Cycle
        ws_all.cell(row=r+2, column=5, value=result[cycle])     # Duration(s)
        cycle = cycle + 1

column_width = [23, 13, 8, 10, 10]
for index, width in enumerate(column_width):
    ws_all.column_dimensions[get_column_letter(index+1)].width = width

wb.save(f'{name_log}_result.xlsx')

print("Success.")
